"""
report_builder.py — Builds all Excel sheets and the Executive Summary PDF.
Keeps all openpyxl / reportlab logic isolated from app.py.
"""

from io import BytesIO
from collections import defaultdict
from datetime import datetime

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# ── Palette ──────────────────────────────────────────────────────────────────
RED_FILL   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
BLUE_FILL  = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
GREY_FILL  = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
HEADER_FILL= PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WHITE_BOLD = Font(bold=True, color="FFFFFF")
DARK_BOLD  = Font(bold=True, color="1F4E79")


def _style_header_row(ws, col_count: int, row: int = 1):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 45)


def _apply_status_fill(ws, status_col: int, total_rows: int, header_row: int = 1):
    for row_num in range(header_row + 1, header_row + total_rows + 1):
        status_cell = ws.cell(row=row_num, column=status_col)
        fill = GREEN_FILL if str(status_cell.value).strip() == "Tested" else RED_FILL
        for c in range(1, ws.max_column + 1):
            ws.cell(row=row_num, column=c).fill = fill
            ws.cell(row=row_num, column=c).border = BORDER


# ── Station Summary builder ───────────────────────────────────────────────────
def build_station_summary(final_rows: list[dict]) -> pd.DataFrame:
    station_data = defaultdict(lambda: {"Tested": 0, "Manual Review Required": 0})
    for row in final_rows:
        st = row["Station"]
        station_data[st][row["Status"]] += 1

    summary_rows = []
    for station, counts in sorted(station_data.items()):
        tested  = counts["Tested"]
        review  = counts["Manual Review Required"]
        summary_rows.append({
            "Station": station,
            "Tested Count": tested,
            "Manual Review Count": review,
            "Total Count": tested + review,
        })
    return pd.DataFrame(summary_rows)


# ── Data Quality engine ───────────────────────────────────────────────────────
def build_data_quality_report(df_raw: pd.DataFrame) -> pd.DataFrame:
    issues = []

    seen = set()
    for idx, row in df_raw.iterrows():
        station  = str(row.get("STATION", "")).strip()
        message  = str(row.get("FAULT MESSAGE", "")).strip()
        time_txt = str(row.get("TIMEDETAILS", "")).strip()

        row_label = f"Row {idx + 5}"  # offset by header rows

        # Missing fields
        if not station or station.lower() == "nan":
            issues.append({"Row": row_label, "Issue Type": "Missing Data",    "Detail": "STATION field is empty",           "Severity": "High"})
        if not message or message.lower() == "nan":
            issues.append({"Row": row_label, "Issue Type": "Missing Data",    "Detail": "FAULT MESSAGE field is empty",      "Severity": "High"})
        if not time_txt or time_txt.lower() == "nan":
            issues.append({"Row": row_label, "Issue Type": "Missing Data",    "Detail": "TIMEDETAILS field is empty",        "Severity": "High"})

        # Duplicate record detection (station + message combo)
        key = (station, message)
        if key in seen:
            issues.append({"Row": row_label, "Issue Type": "Duplicate Record","Detail": f"Station '{station}' with identical FAULT MESSAGE appears more than once", "Severity": "Medium"})
        seen.add(key)

        # Suspicious timestamp detection
        import re
        ts_matches = re.findall(r'\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2}:\d{2}:\d{3}', time_txt)
        from datetime import datetime
        for ts in ts_matches:
            try:
                dt = datetime.strptime(ts[:-4], "%m/%d/%Y %H:%M:%S")
                if dt.year < 2000 or dt.year > 2100:
                    issues.append({"Row": row_label, "Issue Type": "Impossible Timestamp", "Detail": f"Year {dt.year} is outside plausible range in: {ts}", "Severity": "High"})
            except ValueError:
                issues.append({"Row": row_label, "Issue Type": "Invalid Date Format", "Detail": f"Could not parse timestamp: {ts}", "Severity": "Medium"})

    if not issues:
        issues.append({"Row": "—", "Issue Type": "No Issues", "Detail": "Data quality checks passed with no issues found.", "Severity": "—"})

    return pd.DataFrame(issues)


# ── Main Excel package builder ────────────────────────────────────────────────
def build_excel_package(result: dict, df_raw: pd.DataFrame) -> BytesIO:
    """
    Produces a single BytesIO containing all 5 report sheets.
    """
    final_df     = pd.DataFrame(result["final_rows"])
    audit_df     = pd.DataFrame(result["audit_rows"])
    exception_df = pd.DataFrame(result["exception_rows"])
    summary_df   = build_station_summary(result["final_rows"])
    dq_df        = build_data_quality_report(df_raw)

    buf = BytesIO()

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:

        # ── Sheet 1: Final Report ─────────────────────────────────────────
        final_df.to_excel(writer, index=False, sheet_name="Final Report")
        ws = writer.sheets["Final Report"]
        _style_header_row(ws, len(final_df.columns))
        _apply_status_fill(ws, status_col=5, total_rows=len(final_df))
        _auto_width(ws)

        # ── Sheet 2: Exception Report ─────────────────────────────────────
        if exception_df.empty:
            pd.DataFrame([{"Info": "No exceptions — all records classified as Tested."}]).to_excel(
                writer, index=False, sheet_name="Exception Report"
            )
        else:
            exception_df.to_excel(writer, index=False, sheet_name="Exception Report")
            ws_exc = writer.sheets["Exception Report"]
            _style_header_row(ws_exc, len(exception_df.columns))
            for row_num in range(2, len(exception_df) + 2):
                for c in range(1, len(exception_df.columns) + 1):
                    ws_exc.cell(row=row_num, column=c).fill = RED_FILL
                    ws_exc.cell(row=row_num, column=c).border = BORDER
            _auto_width(ws_exc)

        # ── Sheet 3: Audit Report ─────────────────────────────────────────
        audit_df.to_excel(writer, index=False, sheet_name="Audit Report")
        ws_aud = writer.sheets["Audit Report"]
        _style_header_row(ws_aud, len(audit_df.columns))
        _apply_status_fill(ws_aud, status_col=7, total_rows=len(audit_df))
        _auto_width(ws_aud)

        # ── Sheet 4: Station Summary ──────────────────────────────────────
        summary_df.to_excel(writer, index=False, sheet_name="Station Summary")
        ws_sum = writer.sheets["Station Summary"]
        _style_header_row(ws_sum, len(summary_df.columns))
        for row_num in range(2, len(summary_df) + 2):
            for c in range(1, len(summary_df.columns) + 1):
                ws_sum.cell(row=row_num, column=c).fill = BLUE_FILL
                ws_sum.cell(row=row_num, column=c).border = BORDER
        _auto_width(ws_sum)

        # ── Sheet 5: Data Quality Report ──────────────────────────────────
        dq_df.to_excel(writer, index=False, sheet_name="Data Quality Report")
        ws_dq = writer.sheets["Data Quality Report"]
        _style_header_row(ws_dq, len(dq_df.columns))
        severity_fill = {"High": RED_FILL, "Medium": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"), "—": GREY_FILL}
        for row_num in range(2, len(dq_df) + 2):
            sev = str(ws_dq.cell(row=row_num, column=4).value).strip()
            fill = severity_fill.get(sev, GREY_FILL)
            for c in range(1, len(dq_df.columns) + 1):
                ws_dq.cell(row=row_num, column=c).fill = fill
                ws_dq.cell(row=row_num, column=c).border = BORDER
        _auto_width(ws_dq)

    buf.seek(0)
    return buf


# ── PDF Executive Summary ─────────────────────────────────────────────────────
def build_executive_pdf(result: dict) -> BytesIO:
    """
    Generates a clean executive summary PDF using reportlab.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
        )

        buf = BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=A4,
            rightMargin=2 * cm,
            leftMargin=2 * cm,
            topMargin=2 * cm,
            bottomMargin=2 * cm,
        )

        styles = getSampleStyleSheet()
        NAVY   = colors.HexColor("#1F4E79")
        LIGHT  = colors.HexColor("#BDD7EE")
        RED_C  = colors.HexColor("#C00000")
        GREEN_C= colors.HexColor("#375623")

        title_style = ParagraphStyle(
            "title", parent=styles["Title"],
            textColor=NAVY, fontSize=20, spaceAfter=6
        )
        sub_style = ParagraphStyle(
            "sub", parent=styles["Normal"],
            textColor=colors.grey, fontSize=10, spaceAfter=12
        )
        section_style = ParagraphStyle(
            "section", parent=styles["Heading2"],
            textColor=NAVY, fontSize=13, spaceBefore=14, spaceAfter=6
        )
        body_style = ParagraphStyle(
            "body", parent=styles["Normal"],
            fontSize=10, leading=14
        )

        story = []

        story.append(Paragraph("Railway Point Testing Automation", title_style))
        story.append(Paragraph("Executive Summary Report", sub_style))
        story.append(HRFlowable(width="100%", thickness=2, color=NAVY))
        story.append(Spacer(1, 0.4 * cm))

        # Metadata table
        meta = [
            ["Generated On", result["generated_on"]],
            ["Source File",  result.get("filename", "—")],
            ["Rows Processed", str(result["rows_processed"])],
            ["Execution Time", f"{result['execution_time']}s"],
        ]
        meta_table = Table(meta, colWidths=[5 * cm, 11 * cm])
        meta_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), LIGHT),
            ("TEXTCOLOR",  (0, 0), (0, -1), NAVY),
            ("FONTNAME",   (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (1, 0), (1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
            ("BOX",        (0, 0), (-1, -1), 0.5, colors.grey),
            ("INNERGRID",  (0, 0), (-1, -1), 0.25, colors.lightgrey),
            ("LEFTPADDING",(0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 5),
        ]))
        story.append(meta_table)
        story.append(Spacer(1, 0.5 * cm))

        # Key metrics
        story.append(Paragraph("Key Metrics", section_style))

        total = result["tested_count"] + result["review_count"]
        tested_pct = round(result["tested_count"] / total * 100, 1) if total else 0
        review_pct = round(result["review_count"] / total * 100, 1) if total else 0

        metrics = [
            ["Metric", "Count", "Percentage"],
            ["Stations Processed",   str(result["stations_processed"]),  "—"],
            ["Points Processed",     str(result["points_processed"]),     "—"],
            ["Total Records",        str(total),                          "100%"],
            ["Tested",               str(result["tested_count"]),         f"{tested_pct}%"],
            ["Manual Review Required", str(result["review_count"]),       f"{review_pct}%"],
        ]
        metrics_table = Table(metrics, colWidths=[8 * cm, 4 * cm, 4 * cm])
        metrics_table.setStyle(TableStyle([
            ("BACKGROUND",  (0, 0), (-1, 0), NAVY),
            ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
            ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
            ("BOX",         (0, 0), (-1, -1), 0.5, colors.grey),
            ("INNERGRID",   (0, 0), (-1, -1), 0.25, colors.lightgrey),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING",  (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 5),
            ("ALIGN",       (1, 0), (-1, -1), "CENTER"),
        ]))
        story.append(metrics_table)
        story.append(Spacer(1, 0.5 * cm))

        # Notes
        story.append(Paragraph("Classification Policy", section_style))
        story.append(Paragraph(
            "• A point is marked <b>Tested</b> only when at least one valid timestamp "
            "is found within the 06:00–18:00 operational window.<br/>"
            "• Any record with ambiguity, missing data, or parsing failure is marked "
            "<b>Manual Review Required</b>.<br/>"
            "• Records are never automatically classified as 'Not Tested'.",
            body_style
        ))
        story.append(Spacer(1, 0.3 * cm))

        story.append(HRFlowable(width="100%", thickness=1, color=colors.lightgrey))
        story.append(Spacer(1, 0.2 * cm))
        story.append(Paragraph(
            f"<i>Generated by Railway Point Testing Automation System — {result['generated_on']}</i>",
            ParagraphStyle("footer", parent=styles["Normal"], textColor=colors.grey, fontSize=8)
        ))

        doc.build(story)
        buf.seek(0)
        return buf

    except ImportError:
        # If reportlab not installed, return a placeholder
        buf = BytesIO()
        buf.write(b"reportlab not installed. Run: pip install reportlab")
        buf.seek(0)
        return buf
